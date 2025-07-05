#!/usr/bin/env python3
"""
Excel Summary Generator Script

This script creates comprehensive Excel reports from marking results and feedback data.
It generates a single table with one student per row, including:
- Student information
- Total scores and percentages
- Individual module scores with justifications
- Full feedback text (not truncated)
- Test results and quality metrics

Usage:
    python scripts/excel_summary.py --task-name "TASK_NAME" [options]
"""

import argparse
import json
import logging
import pandas as pd
from pathlib import Path
from typing import Dict, List, Optional, Any
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class ExcelSummaryGenerator:
    """Generates comprehensive Excel summaries from marking pipeline results."""
    
    def __init__(self, results_dir: Path, feedback_dir: Path):
        self.results_dir = results_dir
        self.feedback_dir = feedback_dir
    
    def extract_student_name_from_feedback(self, feedback_path: Path) -> Optional[str]:
        """Extract student name from feedback file name or content."""
        try:
            # Try to extract from filename first
            filename = feedback_path.stem
            # Pattern: Name_Name_zID_task_feedback
            parts = filename.split('_')
            if len(parts) >= 3:
                # Find the student ID (starts with 'z')
                student_id_idx = None
                for i, part in enumerate(parts):
                    if part.startswith('z') and part[1:].isdigit():
                        student_id_idx = i
                        break
                
                if student_id_idx and student_id_idx > 0:
                    # Name parts are before the student ID
                    name_parts = parts[:student_id_idx]
                    return ' '.join(name_parts).replace('_', ' ')
            
            # Fallback: try to extract from file content
            content = feedback_path.read_text(encoding='utf-8')
            match = re.search(r'Student:\s*([^(]+)', content)
            if match:
                return match.group(1).strip()
                
        except Exception as e:
            logger.warning(f"Could not extract student name from {feedback_path}: {e}")
        
        return None
    
    def load_scores_data(self, task_name: str) -> Dict[str, Any]:
        """Load the scores summary JSON file."""
        scores_file = self.results_dir / task_name / "scores_summary.json"
        
        if not scores_file.exists():
            logger.warning(f"Scores summary file not found: {scores_file}")
            return {"students": []}
        
        try:
            with open(scores_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"Failed to load scores data: {e}")
            return {"students": []}
    
    def load_test_results(self, task_name: str) -> Dict[str, Dict]:
        """Load test results for all students."""
        test_results = {}
        task_results_dir = self.results_dir / task_name
        
        if not task_results_dir.exists():
            return test_results
        
        for test_file in task_results_dir.glob("*_test_results.json"):
            try:
                with open(test_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    student_id = self._extract_student_id_from_filename(test_file.name)
                    if student_id:
                        test_results[student_id] = data
            except Exception as e:
                logger.warning(f"Failed to load test results from {test_file}: {e}")
        
        return test_results
    
    def load_quality_results(self, task_name: str) -> Dict[str, Dict]:
        """Load quality results for all students."""
        quality_results = {}
        task_results_dir = self.results_dir / task_name
        
        if not task_results_dir.exists():
            return quality_results
        
        for quality_file in task_results_dir.glob("*_quality_results.json"):
            try:
                with open(quality_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    student_id = self._extract_student_id_from_filename(quality_file.name)
                    if student_id:
                        quality_results[student_id] = data
            except Exception as e:
                logger.warning(f"Failed to load quality results from {quality_file}: {e}")
        
        return quality_results
    
    def load_feedback_data(self, task_name: str) -> Dict[str, Dict]:
        """Load feedback data for all students."""
        feedback_data = {}
        task_feedback_dir = self.feedback_dir / task_name
        
        if not task_feedback_dir.exists():
            logger.warning(f"Feedback directory not found: {task_feedback_dir}")
            return feedback_data
        
        for feedback_file in task_feedback_dir.glob("*.md"):
            try:
                student_id = self._extract_student_id_from_filename(feedback_file.name)
                if student_id:
                    content = feedback_file.read_text(encoding='utf-8')
                    student_name = self.extract_student_name_from_feedback(feedback_file)
                    
                    feedback_data[student_id] = {
                        'student_name': student_name,
                        'feedback_content': content,
                        'feedback_length': len(content),
                        'feedback_file': str(feedback_file)
                    }
            except Exception as e:
                logger.warning(f"Failed to load feedback from {feedback_file}: {e}")
        
        return feedback_data
    
    def _extract_student_id_from_filename(self, filename: str) -> Optional[str]:
        """Extract student ID (zXXXXXXX) from filename."""
        match = re.search(r'(z\d+)', filename)
        return match.group(1) if match else None
    
    def calculate_test_summary(self, test_data: Dict) -> Dict[str, Any]:
        """Calculate summary statistics from test results."""
        if not test_data:
            return {
                'total_tests': 0,
                'passed_tests': 0,
                'failed_tests': 0,
                'test_pass_rate': 0.0,
                'test_score': 0.0
            }
        
        total_tests = 0
        passed_tests = 0
        
        # Handle different test result structures
        if 'test_results' in test_data:
            results = test_data['test_results']
            if isinstance(results, list):
                total_tests = len(results)
                passed_tests = sum(1 for r in results if r.get('passed', False))
            elif isinstance(results, dict):
                total_tests = len(results)
                passed_tests = sum(1 for r in results.values() if r.get('passed', False))
        
        pass_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0
        
        return {
            'total_tests': total_tests,
            'passed_tests': passed_tests,
            'failed_tests': total_tests - passed_tests,
            'test_pass_rate': round(pass_rate, 2),
            'test_score': round(pass_rate, 2)
        }
    
    def calculate_quality_summary(self, quality_data: Dict) -> Dict[str, Any]:
        """Calculate summary from quality metrics."""
        if not quality_data:
            return {
                'code_quality_score': 0.0,
                'complexity_score': 0.0,
                'documentation_score': 0.0,
                'style_issues': 0
            }
        
        summary = {
            'code_quality_score': quality_data.get('overall_score', 0.0),
            'complexity_score': quality_data.get('complexity_score', 0.0),
            'documentation_score': quality_data.get('documentation_score', 0.0),
            'style_issues': quality_data.get('style_issues_count', 0)
        }
        
        return summary
    
    def check_submission_quality(self, student_data: Dict, test_summary: Dict, quality_summary: Dict) -> Dict[str, Any]:
        """Check for potential submission problems and generate warnings."""
        warnings = []
        warning_flags = {
            'low_total_score': False,
            'many_zero_scores': False,
            'no_test_results': False,
            'extraction_errors': False,
            'missing_feedback': False
        }
        
        # Check 1: Very low total score (less than 10% of maximum)
        total_score = student_data.get('total_score', 0)
        max_total_score = student_data.get('max_total_score', 1)
        percentage = (total_score / max_total_score * 100) if max_total_score > 0 else 0
        
        if percentage < 10:
            warnings.append(f"Very low total score ({percentage:.1f}%)")
            warning_flags['low_total_score'] = True
        
        # Check 2: Too many zero scores in modules
        module_scores = student_data.get('module_scores', [])
        zero_count = sum(1 for module in module_scores if module.get('score', 0) == 0)
        total_modules = len(module_scores)
        
        if total_modules > 0:
            zero_percentage = (zero_count / total_modules * 100)
            if zero_percentage >= 50:  # 50% or more modules got zero
                warnings.append(f"{zero_count}/{total_modules} modules scored zero ({zero_percentage:.1f}%)")
                warning_flags['many_zero_scores'] = True
        
        # Check 3: No test results or all tests failed
        if test_summary['total_tests'] == 0:
            warnings.append("No test results found")
            warning_flags['no_test_results'] = True
        elif test_summary['passed_tests'] == 0 and test_summary['total_tests'] > 0:
            warnings.append(f"All {test_summary['total_tests']} tests failed")
            warning_flags['no_test_results'] = True
        
        # Check 4: Many extraction errors
        extraction_errors = len(student_data.get('extraction_errors', []))
        if extraction_errors >= 3:
            warnings.append(f"{extraction_errors} score extraction errors")
            warning_flags['extraction_errors'] = True
        
        # Check 5: Missing feedback
        if not student_data.get('student_name') or student_data.get('student_name') == 'Unknown':
            warnings.append("Missing student information")
            warning_flags['missing_feedback'] = True
        
        # Generate overall assessment
        has_problems = any(warning_flags.values())
        
        if has_problems:
            problem_summary = "⚠️ POTENTIAL SUBMISSION PROBLEMS - Please check original submission file"
            detailed_warning = "; ".join(warnings)
        else:
            problem_summary = "✅ No obvious submission problems detected"
            detailed_warning = ""
        
        return {
            'has_problems': has_problems,
            'problem_summary': problem_summary,
            'detailed_warnings': detailed_warning,
            'warning_count': len(warnings),
            'warning_flags': warning_flags,
            'individual_warnings': warnings
        }
    
    def get_all_module_ids(self, scores_data: Dict) -> List[str]:
        """Get all unique module IDs from the scores data."""
        module_ids = set()
        for student_data in scores_data.get('students', []):
            for module_score in student_data.get('module_scores', []):
                module_ids.add(module_score['module_id'])
        return sorted(list(module_ids))
    
    def generate_comprehensive_excel(self, task_name: str, output_path: Optional[Path] = None) -> Path:
        """Generate a comprehensive Excel file with one student per row."""
        logger.info(f"Generating comprehensive Excel for task: {task_name}")
        
        # Load all data sources
        scores_data = self.load_scores_data(task_name)
        test_results = self.load_test_results(task_name)
        quality_results = self.load_quality_results(task_name)
        feedback_data = self.load_feedback_data(task_name)
        
        # Get all module IDs to create consistent columns
        all_module_ids = self.get_all_module_ids(scores_data)
        
        # Prepare data for Excel
        excel_data = []
        
        for student_data in scores_data.get('students', []):
            student_id = student_data['student_id']
            
            # Basic student info
            row = {
                'Student_ID': student_id,
                'Student_Name': student_data.get('student_name') or 
                              feedback_data.get(student_id, {}).get('student_name', 'Unknown'),
                'Task_Name': task_name,
                'Submission_Date': datetime.now().strftime('%Y-%m-%d'),
            }
            
            # Overall scoring information
            row.update({
                'Total_Score': student_data.get('total_score', 0),
                'Max_Total_Score': student_data.get('max_total_score', 0),
                'Percentage_Score': round(student_data.get('percentage', 0), 2),
                'Module_Count': len(student_data.get('module_scores', [])),
                'Extraction_Errors': len(student_data.get('extraction_errors', []))
            })
            
            # Test results summary
            test_summary = self.calculate_test_summary(test_results.get(student_id, {}))
            row.update({
                'Test_Total': test_summary['total_tests'],
                'Test_Passed': test_summary['passed_tests'],
                'Test_Failed': test_summary['failed_tests'],
                'Test_Pass_Rate': test_summary['test_pass_rate'],
                'Test_Score': test_summary['test_score']
            })
            
            # Quality metrics summary
            quality_summary = self.calculate_quality_summary(quality_results.get(student_id, {}))
            row.update({
                'Quality_Code_Score': quality_summary['code_quality_score'],
                'Quality_Complexity': quality_summary['complexity_score'],
                'Quality_Documentation': quality_summary['documentation_score'],
                'Quality_Style_Issues': quality_summary['style_issues']
            })
            
            # Create a lookup for module scores
            module_scores_lookup = {}
            for module_score in student_data.get('module_scores', []):
                module_scores_lookup[module_score['module_id']] = module_score
            
            # Add module scores and justifications
            for module_id in all_module_ids:
                if module_id in module_scores_lookup:
                    module_score = module_scores_lookup[module_id]
                    row[f'Module_{module_id}_Score'] = module_score['score']
                    row[f'Module_{module_id}_Max'] = module_score['max_score']
                    row[f'Module_{module_id}_Percentage'] = round(
                        (module_score['score'] / module_score['max_score'] * 100) 
                        if module_score['max_score'] > 0 else 0, 2
                    )
                    row[f'Module_{module_id}_Success'] = module_score['success']
                    row[f'Module_{module_id}_Justification'] = module_score['justification']
                else:
                    # Module not found for this student
                    row[f'Module_{module_id}_Score'] = 0
                    row[f'Module_{module_id}_Max'] = 0
                    row[f'Module_{module_id}_Percentage'] = 0
                    row[f'Module_{module_id}_Success'] = False
                    row[f'Module_{module_id}_Justification'] = 'Module not evaluated'
            
            # Full feedback content
            feedback_info = feedback_data.get(student_id, {})
            row.update({
                'Has_Feedback': bool(feedback_info),
                'Feedback_Length': feedback_info.get('feedback_length', 0),
                'Feedback_File_Path': feedback_info.get('feedback_file', ''),
                'Full_Feedback_Text': feedback_info.get('feedback_content', 'No feedback available')
            })
            
            # Submission quality check
            quality_check = self.check_submission_quality(student_data, test_summary, quality_summary)
            row.update({
                'Has_Problems': quality_check['has_problems'],
                'Problem_Summary': quality_check['problem_summary'],
                'Detailed_Warnings': quality_check['detailed_warnings'],
                'Warning_Count': quality_check['warning_count'],
                'Warning_Flags': json.dumps(quality_check['warning_flags']),
                'Individual_Warnings': json.dumps(quality_check['individual_warnings'])
            })
            
            excel_data.append(row)
        
        # Create DataFrame
        df = pd.DataFrame(excel_data)
        
        # Generate output path if not provided
        if output_path is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            safe_task_name = task_name.replace(' ', '_').replace('(', '').replace(')', '')
            output_path = Path(f"comprehensive_{safe_task_name}_{timestamp}.xlsx")
        
        # Write to Excel with formatting
        self._write_formatted_excel(df, output_path, task_name, all_module_ids)
        
        logger.info(f"Comprehensive Excel saved to: {output_path}")
        return output_path
    
    def _write_formatted_excel(self, df: pd.DataFrame, output_path: Path, task_name: str, module_ids: List[str]):
        """Write DataFrame to Excel with proper formatting."""
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Student Summary"
        
        # Write data to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Problem highlighting styles
        problem_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # Light red
        warning_font = Font(bold=True, color="CC0000")  # Dark red
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Find the "Has_Problems" column index for conditional formatting
        has_problems_col = None
        problem_summary_col = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == 'Has_Problems':
                has_problems_col = idx
            elif cell.value == 'Problem_Summary':
                problem_summary_col = idx
        
        # Auto-adjust column widths and format cells
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            column_index = column[0].column
            
            for row_idx, cell in enumerate(column, 1):
                cell.border = border
                
                # Check if this row has problems (skip header row)
                has_problems = False
                if row_idx > 1 and has_problems_col:
                    problem_cell = ws.cell(row=row_idx, column=has_problems_col)
                    has_problems = problem_cell.value is True
                
                # Apply problem highlighting
                if has_problems and row_idx > 1:
                    cell.fill = problem_fill
                    # Make problem summary column text bold and red
                    if column_index == problem_summary_col:
                        cell.font = warning_font
                
                # Handle different column types
                if cell.value is not None:
                    # Special formatting for problem summary column
                    if column_index == problem_summary_col:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                        ws.column_dimensions[column_letter].width = 60
                        continue
                    # Wrap text for justification and feedback columns
                    elif any(keyword in str(cell.column_letter) for keyword in ['Justification', 'Feedback_Text', 'Detailed_Warnings']):
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                        # Set a reasonable width for text columns
                        ws.column_dimensions[column_letter].width = 50
                        continue
                    else:
                        cell.alignment = Alignment(vertical='center')
                    
                    # Calculate column width
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            
            # Set column width (with limits)
            if not any(keyword in column_letter for keyword in ['Justification', 'Feedback_Text', 'Problem_Summary', 'Detailed_Warnings']):
                adjusted_width = min(max_length + 2, 30)  # Cap at 30 characters
                ws.column_dimensions[column_letter].width = max(adjusted_width, 10)
        
        # Set row heights for better readability, extra height for problem rows
        for row_num in range(2, ws.max_row + 1):  # Skip header
            has_problems = False
            if has_problems_col:
                problem_cell = ws.cell(row=row_num, column=has_problems_col)
                has_problems = problem_cell.value is True
            
            # Set higher row height for problem rows
            ws.row_dimensions[row_num].height = 45 if has_problems else 30
        
        # Freeze panes (freeze first row and first few columns)
        ws.freeze_panes = 'E2'  # Freeze first 4 columns and header row
        
        # Add a summary sheet
        self._add_summary_sheet(wb, df, task_name, module_ids)
        
        # Save workbook
        wb.save(output_path)
    
    def _add_summary_sheet(self, wb: openpyxl.Workbook, df: pd.DataFrame, task_name: str, module_ids: List[str]):
        """Add a summary statistics sheet to the workbook."""
        ws_summary = wb.create_sheet("Summary Statistics")
        
        # Task information
        ws_summary['A1'] = "Task Summary"
        ws_summary['A1'].font = Font(bold=True, size=14)
        
        ws_summary['A3'] = "Task Name:"
        ws_summary['B3'] = task_name
        ws_summary['A4'] = "Total Students:"
        ws_summary['B4'] = len(df)
        ws_summary['A5'] = "Report Generated:"
        ws_summary['B5'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Problem detection summary
        ws_summary['A7'] = "Submission Quality Check"
        ws_summary['A7'].font = Font(bold=True, size=12, color="CC0000")
        
        if 'Has_Problems' in df.columns:
            problem_count = len(df[df['Has_Problems'] == True])
            ws_summary['A9'] = "Students with Problems:"
            ws_summary['B9'] = f"{problem_count}/{len(df)}"
            ws_summary['B9'].font = Font(bold=True, color="CC0000" if problem_count > 0 else "008000")
            
            if problem_count > 0:
                ws_summary['A10'] = "⚠️ CHECK THESE SUBMISSIONS:"
                ws_summary['A10'].font = Font(bold=True, color="CC0000")
                
                # List problematic students
                problem_students = df[df['Has_Problems'] == True]
                row = 11
                for _, student in problem_students.iterrows():
                    student_info = f"{student.get('Student_Name', 'Unknown')} ({student.get('Student_ID', 'Unknown')})"
                    problem_summary = student.get('Problem_Summary', 'Unknown issues')
                    
                    ws_summary[f'A{row}'] = student_info
                    ws_summary[f'B{row}'] = problem_summary
                    ws_summary[f'A{row}'].font = Font(bold=True)
                    ws_summary[f'B{row}'].font = Font(color="CC0000")
                    row += 1
                
                # Add spacing
                row += 1
            else:
                ws_summary['A10'] = "✅ All submissions appear normal"
                ws_summary['A10'].font = Font(color="008000")
                row = 12
        else:
            row = 9
        
        # Score statistics
        ws_summary[f'A{row}'] = "Score Statistics"
        ws_summary[f'A{row}'].font = Font(bold=True, size=12)
        row += 2
        
        if 'Total_Score' in df.columns:
            ws_summary[f'A{row}'] = "Average Total Score:"
            ws_summary[f'B{row}'] = round(df['Total_Score'].mean(), 2)
            row += 1
            ws_summary[f'A{row}'] = "Median Total Score:"
            ws_summary[f'B{row}'] = round(df['Total_Score'].median(), 2)
            row += 1
            ws_summary[f'A{row}'] = "Min Total Score:"
            ws_summary[f'B{row}'] = df['Total_Score'].min()
            row += 1
            ws_summary[f'A{row}'] = "Max Total Score:"
            ws_summary[f'B{row}'] = df['Total_Score'].max()
            row += 2
        
        if 'Percentage_Score' in df.columns:
            ws_summary[f'A{row}'] = "Average Percentage:"
            ws_summary[f'B{row}'] = f"{round(df['Percentage_Score'].mean(), 2)}%"
            row += 1
            ws_summary[f'A{row}'] = "Students Above 70%:"
            ws_summary[f'B{row}'] = len(df[df['Percentage_Score'] >= 70])
            row += 1
            ws_summary[f'A{row}'] = "Students Below 50%:"
            ws_summary[f'B{row}'] = len(df[df['Percentage_Score'] < 50])
            row += 1
            ws_summary[f'A{row}'] = "Students Below 10%:"
            ws_summary[f'B{row}'] = len(df[df['Percentage_Score'] < 10])
            ws_summary[f'B{row}'].font = Font(bold=True, color="CC0000" if len(df[df['Percentage_Score'] < 10]) > 0 else "000000")
            row += 2
        
        # Module performance summary
        ws_summary[f'A{row}'] = "Module Performance"
        ws_summary[f'A{row}'].font = Font(bold=True, size=12)
        row += 2
        
        for module_id in module_ids[:10]:  # Show top 10 modules
            score_col = f'Module_{module_id}_Score'
            max_col = f'Module_{module_id}_Max'
            
            if score_col in df.columns and max_col in df.columns:
                avg_score = df[score_col].mean()
                max_possible = df[max_col].iloc[0] if len(df) > 0 else 0
                avg_percentage = (avg_score / max_possible * 100) if max_possible > 0 else 0
                
                ws_summary[f'A{row}'] = module_id
                ws_summary[f'B{row}'] = f"{round(avg_score, 1)}/{max_possible} ({round(avg_percentage, 1)}%)"
                row += 1
        
        # Auto-adjust column widths for summary sheet
        for column in ws_summary.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value is not None:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = min(max_length + 2, 80)  # Increased max width for problem descriptions
            ws_summary.column_dimensions[column_letter].width = max(adjusted_width, 15)


def main():
    parser = argparse.ArgumentParser(description="Generate comprehensive Excel summaries from marking results")
    parser.add_argument("--task-name", type=str, required=True, 
                       help="Name of the task to generate summary for")
    parser.add_argument("--results-dir", type=Path, default=Path("results"),
                       help="Directory containing results data")
    parser.add_argument("--feedback-dir", type=Path, default=Path("feedback"),
                       help="Directory containing feedback files")
    parser.add_argument("--output-file", type=Path,
                       help="Output Excel file path (optional)")
    parser.add_argument("--output-dir", type=Path, default=Path("reports"),
                       help="Directory to save Excel report if no specific file specified")
    
    args = parser.parse_args()
    
    # Validate directories
    if not args.results_dir.exists():
        logger.error(f"Results directory not found: {args.results_dir}")
        return 1
    
    if not args.feedback_dir.exists():
        logger.warning(f"Feedback directory not found: {args.feedback_dir}")
    
    # Create output directory if needed
    if args.output_file:
        args.output_file.parent.mkdir(parents=True, exist_ok=True)
        output_path = args.output_file
    else:
        args.output_dir.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_task_name = args.task_name.replace(' ', '_').replace('(', '').replace(')', '')
        output_path = args.output_dir / f"comprehensive_{safe_task_name}_{timestamp}.xlsx"
    
    # Initialize generator
    generator = ExcelSummaryGenerator(args.results_dir, args.feedback_dir)
    
    try:
        result_path = generator.generate_comprehensive_excel(args.task_name, output_path)
        print(f"Comprehensive Excel report generated: {result_path}")
        return 0
        
    except Exception as e:
        logger.error(f"Failed to generate Excel report: {e}", exc_info=True)
        return 1


if __name__ == "__main__":
    exit(main()) 